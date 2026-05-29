import numpy as np
import matplotlib.pyplot as plt

from openpyxl import load_workbook

file_path = r"C:\Users\Jaleal\Desktop\EAE 130B\Zephyr One Database.xlsx"
wb = load_workbook(file_path)
ws = wb["Database"]

N_en = ws["B14"].value
T = ws["B12"].value
N_z = ws["B18"].value

tc_root = ws["B119"].value
wing_sweep_qc = ws["B118"].value
W_0 = ws["B11"].value
W_zf = ws["B29"].value

W_engine_mounts = 0.013*N_en**0.795*T**0.579*N_z
W_firewall = ws["B38"].value
W_engine = ws["B31"].value
W_engine_section = 0.01*W_engine**0.717*N_en*N_z

K_vg = ws["B170"].value # non-variable geometry
L_d = ws["B159"].value # ft
K_d = ws["B171"].value # more elliptical shape
L_s = L_d / 3 # ft
D_e = ws["B154"].value # ft
W_air_induction_system = 13.29*K_vg*L_d**0.643*K_d**0.182*N_en**1.498*(L_s/L_d)**-0.373*D_e

L_tp = ws["B160"].value # ft
W_tailpipe = 3.5*D_e*L_tp*N_en

L_sh = ws["B161"].value # ft, guess
W_engine_cooling = 4.55*D_e*L_sh*N_en

W_oil_cooling = 37.82*N_en**1.023

T_e = ws["B12"].value # lbf
W_starter = 0.025*T_e**0.760*N_en**0.72

# Constants and Inputs
K_cb = ws["B172"].value
K_mc = ws["B174"].value
K_tpg = ws["B173"].value
K_vsh = ws["B126"].value
L_a = ws["B162"].value
L_ec = ws["B163"].value
L_m = ws["B164"].value
L_n = ws["B165"].value
M = ws["B102"].value
N_c = ws["B15"].value
N_ci = ws["B15"].value
N_en = ws["B14"].value
N_gen = ws["B176"].value
N_nw = ws["B177"].value
N_l = ws["B178"].value
N_s = ws["B179"].value
N_t = ws["B180"].value
N_u = ws["B181"].value
R_kva = ws["B182"].value
S_cs = ws["B116"].value
W_dg = ws["B37"].value
W_l = ws["B35"].value
W_uav = ws["B36"].value
taper_ratio_wing = ws["B120"].value

# Empennage
F_w = ws["B153"].value # ft
B_h = ws["B130"].value # ft
S_ht = ws["B128"].value # ft^2
W_horizontal_tail = 3.316*(1+F_w/B_h)**-2.0*(W_dg*N_z/1000)**0.260*S_ht**0.806

K_rht = ws["B136"].value # non-rolling horizontal tail
H_t = ws["B135"].value # horizontal tail not above fuselage
H_v = ws["B146"].value # vertical tail needs to be moved above fuselage
S_vt = ws["B138"].value # ft^2
M = ws["B102"].value
L_t = ws["B134"].value # ft
S_r = ws["B140"].value # ft^2
A_vt = ws["B142"].value
taper_ratio_tail = ws["B144"].value
tail_sweep_qc = ws["B145"].value # estimate
W_vertical_tail = 0.452*K_rht*(1+H_t/H_v)**0.5*(W_dg*N_z)**0.488*S_vt**0.718*M**0.341*(
    L_t)**-1.0*(1+S_r/S_vt)**0.348*A_vt**0.223*(1+taper_ratio_tail)**0.25*(np.cos(np.deg2rad(tail_sweep_qc)))**-0.323

# Fuselage
K_dwf = ws["B125"].value
L = ws["B149"].value # ft
D = ws["B150"].value # ft
W = ws["B151"].value # ft
W_fuselage = 0.499*K_dwf*W_dg**0.35*N_z**0.25*L**0.5*D**0.849*W**0.685

# Calculations
W_crew = ws["B32"].value

W_mainlandinggear = K_cb * K_tpg * (W_l * N_l)**0.25 * L_m**0.973
W_noselanding = (W_l * N_l)**0.29 * L_n**0.5 * N_nw**0.525

W_landinggear = 0.033*(41562.51 + 17604)
W_mainlandinggear = 0.85*W_landinggear
W_noselanding = 0.15*W_landinggear

W_flightcontrols = 36.28 * M**0.003 * S_cs**0.489 * N_s**0.484 * N_c**0.127
W_enginecontrols = 10.5 * N_en**1.008 * L_ec**0.222

W_instruments = 8 + 36.37 * N_en**0.676 * N_t**0.237 + 26.4 * (1 + N_ci)**1.356
W_hydraulics = 37.23 * K_vsh * N_u**0.664
W_electrical = 172.2 * K_mc * R_kva**0.152 * N_c**0.1 * L_a**0.1 * N_gen**0.091
W_avionics = 2.117 * W_uav**0.933
W_furnishings = 217.6 * N_c
W_airconditioning = 201.6 * ((W_uav + 200 * N_c) / 100)**0.735
W_handlinggear = 3.2e-4 * W_dg

# Resulting Sum
W_adrian = (
    W_crew + W_mainlandinggear + W_noselanding + W_flightcontrols + W_enginecontrols + 
    W_instruments + W_hydraulics + W_electrical + W_avionics + 
    W_furnishings + W_airconditioning + W_handlinggear + 16000
)

# COST MODEL
Q = ws["B51"].value
CPI = ws["B52"].value
V = ws["B94"].value  # velocity

R = [115, 118, 98, 108, 1, 1, 1]
coeff = [4.86, 5.99, 7.37, 0.076, 91.3, 2498, 22.1]
W_exp = [0.777, 0.777, 0.82, 0, 0.630, 0.325, 0.921]
V_exp = [0.894, 0.696, 0.484, 0, 1.3, 0.822, 0.621]
Q_exp = [0.163, 0.263, 0.641, 0, 0, 0, 0.799]

rdte = [0, 4, 5]
prod = [1, 2, 3, 6]

C_engine = ws["B56"].value # Cost of P135 P&W 100

# Avionics cost
CPI_av = ws["B53"].value
avionics_cost = 2000 * W_avionics * CPI_av

W_fuel = ws["B30"].value

## FUNCTION ##
def compute_outputs(A, S_w):

    b = np.sqrt(A * S_w)

    W_wing = 4.22*S_w + 1.642*10**-6*(N_z*b**3*np.sqrt(W_0*W_zf)*(1+2*taper_ratio_wing))/(S_w*tc_root*(np.cos(np.deg2rad(wing_sweep_qc)))**2*(1+taper_ratio_wing))


    W_total = (
        W_engine_mounts + W_firewall + W_engine_section + W_engine +
        W_air_induction_system + W_tailpipe + W_engine_cooling +
        W_oil_cooling + W_starter +
        W_wing + W_horizontal_tail + W_vertical_tail +
        W_fuselage + W_adrian
    )

    WS = (W_total + W_fuel)/ S_w

    C_rdte = []
    for i in rdte:
        C_rdte.append(coeff[i]*(W_total**W_exp[i])*(V**V_exp[i])*(Q**Q_exp[i])*R[i]*CPI)

    C_prod = []
    for i in prod:
        if i == 3:
            y = 2
            mfg_hours = coeff[y]*(W_total**W_exp[y])*(V**V_exp[y])*(Q**Q_exp[y])
            C_prod.append(coeff[i]*R[i]*mfg_hours*CPI)
        else:
            C_prod.append(coeff[i]*(W_total**W_exp[i])*(V**V_exp[i])*(Q**Q_exp[i])*R[i]*CPI)

    C_flyaway = (sum(C_prod)/Q) + C_engine + avionics_cost

    return C_flyaway, WS


# GRID
AR_vals = np.linspace(2, 7, 25)
S_vals = np.linspace(300, 900, 25)

C_grid = np.zeros((len(AR_vals), len(S_vals)))
WS_grid = np.zeros((len(AR_vals), len(S_vals)))

for i, A in enumerate(AR_vals):
    for j, S in enumerate(S_vals):
        C, WS = compute_outputs(A, S)
        C_grid[i, j] = C
        WS_grid[i, j] = WS


## PLOT ##
plt.figure(figsize=(10, 8))

# Constant AR curves
for i in range(len(AR_vals)):
    plt.plot(C_grid[i, :], WS_grid[i, :], 'k')

# Constant S curves
for j in range(len(S_vals)):
    plt.plot(C_grid[:, j], WS_grid[:, j], 'k')

# Labels
x_offset = 0.01 * (C_grid.max() - C_grid.min())
y_offset = 0.01 * (WS_grid.max() - WS_grid.min())

# AR labels 
for i in range(0, len(AR_vals), 6):
    plt.text(C_grid[i, -1] + x_offset, WS_grid[i, -1],
             f'AR={AR_vals[i]:.1f}',
             fontsize=10, ha='left', va='center')

# S labels 
for j in range(0, len(S_vals), 6):
    plt.text(C_grid[0, j], WS_grid[0, j] - y_offset,
             f'S={S_vals[j]:.0f}',
             fontsize=10, ha='center', va='top')

# DESIGN POINT
A_point = 3.3
S_point = 485

C_point, WS_point = compute_outputs(A_point, S_point)

plt.scatter(C_point, WS_point, color='red', s=80, zorder=5)
plt.annotate(
    f'Current: AR={A_point:.2f}, S={S_point:.2f}',
    (C_point, WS_point),
    textcoords="offset points",
    xytext=(15, 5),             
    ha='left',
    va='center',               
    fontsize=11,
    color='red',
    bbox=dict(
        boxstyle='round,pad=0.3', 
        fc='white',            
        ec='none',             
        alpha=0.8              
    )
)

A_point_new = 3.02
S_point_new = 530

C_point_new, WS_point_new = compute_outputs(A_point_new, S_point_new)

plt.scatter(C_point_new, WS_point_new, color='blue', s=80, zorder=5)
plt.annotate(
    f'New: AR={A_point_new:.2f}, S={S_point_new:.2f}',
    (C_point_new, WS_point_new),
    textcoords="offset points",
    xytext=(15, 5),             
    ha='left',
    va='center',               
    fontsize=11,
    color='blue',
    bbox=dict(
        boxstyle='round,pad=0.3', 
        fc='white',            
        ec='none',             
        alpha=0.8              
    )
)

plt.xlabel('Flyaway Cost ($)', fontsize=14)
plt.ylabel('W/S (lb/ft²)', fontsize=14)
plt.title('Cost vs Wing Loading Carpet Plot', fontsize=16)

plt.grid(True, linestyle='--', alpha=0.5)
plt.show()

wb.save(r"C:\Users\Jaleal\Desktop\EAE 130B\Zephyr One Database.xlsx")