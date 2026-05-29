# SECTION 1
import numpy as np
import math
import pandas as pd
import matplotlib.pyplot as plt

from openpyxl import load_workbook

file_path = r"C:\Users\Jaleal\Desktop\EAE 130B\Zephyr One Database.xlsx"
wb = load_workbook(file_path)
ws = wb["Database"]

# SECTION 2
def compute_stall_boundary(V_min, V_max, n_points, rho, CL_max, MTOW, S_ref):
    print("Computing stall boundary...")
    stall_coeff = 0.5 * rho * CL_max / (MTOW / S_ref)
    print(f"W/S: {MTOW / S_ref:.3f} lb/ft^2")
    print(f"Stall coefficient: {stall_coeff:.6f}")

    V = np.linspace(V_min, V_max, n_points)
    n_Stall = stall_coeff * V**2

    return V, n_Stall, stall_coeff

# Parameters
V_min = 0          # ft/s, minimum speed
n_points = 100     # number of points in the speed range
rho = ws["B60"].value    # slug/ft^3, air density at conditions of sea level
CL_max = ws["B68"].value    # maximum lift coefficient, takeoff (1.623 for landing)
#MTOW = 62473       # lb, maximum takeoff weight (MAX) includes empty weight + full fuel + payload
MTOW = ws["B28"].value      # lb, maximum takeoff weight (MIN) includes empty weight + reserve fuel
S_ref = ws["B8"].value  # ft^2, reference wing area

#V_max calculation
def calculate_Vmax(T_max, MTOW, S_ref, rho, CD0, e, AR):
    """
    Simplest Vmax calculation using iteration
    
    Returns: V_max in ft/s
    """
    # Initial guess (start high)
    V = 500  # ft/s
    
    # Iterate until convergence
    for i in range(20):  # 20 iterations is plenty
        # Calculate CL from lift = weight
        CL = MTOW / (0.5 * rho * V**2 * S_ref)
        
        # Calculate CD from drag polar
        CD = CD0 + (CL**2) / (np.pi * e * AR)
        
        # Calculate drag force
        Drag = 0.5 * rho * V**2 * S_ref * CD
        
        # Adjust velocity based on thrust - drag imbalance
        if Drag > T_max:
            V = V * 0.98  # Too fast, slow down
        else:
            V = V * 1.02  # Too slow, speed up
    
    return V

# V_max parameters
T_max = ws["B12"].value      # lb, maximum sea level thrust
CD0 = ws["B74"].value    # zero-lift drag coefficient
e = ws["B88"].value          # Oswald efficiency factor
AR = ws["B10"].value       # aspect ratio

V_max = calculate_Vmax(T_max, MTOW, S_ref, rho, CD0, e, AR)
print(f"V_max = {V_max:.1f} ft/s")

Boundary_V, Boundary_n_Stall, stall_coeff = compute_stall_boundary(V_min, V_max, n_points, rho, CL_max, MTOW, S_ref)

V = np.linspace(V_min, V_max, n_points)

plt.figure(figsize=(16,9))
plt.title('Maneuvering Envelope')
plt.xlabel("V (ft/s)")
plt.ylabel("n (-)")
plt.plot(Boundary_V, Boundary_n_Stall, label='Stall', linestyle='-', linewidth=2, marker=None, markersize=8)
plt.plot(Boundary_V, -Boundary_n_Stall, label='Stall', linestyle='-', linewidth=2, marker=None, markersize=8)
plt.grid(True)
plt.legend(loc='best')
plt.show()

# SECTION 3

# For positive limit load factor, Table 14.2 has typical values of 6.5 to 9, with RFP calling out n_pos > 7g
n_pos_limit = ws["B18"].value
n_pos_limit = np.ones(100)*n_pos_limit

# For negative limit load factor, Table 14.2 has typical values of -3 to -6, RFP doesn't call out min values specifically but chosen as -3
n_neg_limit = ws["B19"].value
n_neg_limit = np.ones(100)*n_neg_limit
print(f"Negative limit load factor: {n_neg_limit[0]:.2f}")

plt.figure(figsize=(16,9))
plt.title('Maneuvering Envelope')
plt.xlabel("V (ft/s)")
plt.ylabel("n (-)")
plt.plot(Boundary_V, Boundary_n_Stall, label='Stall', linestyle='-', linewidth=2, marker=None, markersize=8)
plt.plot(Boundary_V, -Boundary_n_Stall, label='Stall', linestyle='-', linewidth=2, marker=None, markersize=8)
plt.plot(V,n_pos_limit, label='Limit Load Pos', linestyle='-', linewidth=2, marker=None, markersize=8)
plt.plot(V,n_neg_limit, label='Limit Load Neg', linestyle='-', linewidth=2, marker=None, markersize=8)
plt.grid(True)
plt.legend(loc='best')
plt.show()


# SECTION 4
def compute_intersection_velocity(stall_coeff, n_limit):
    """
    Solve stall_coeff * V^2 = |n_limit|
    """
    int_V = np.sqrt(abs(n_limit) / stall_coeff)
    print(f"Intersection velocity for n_limit={n_limit:.2f} is {int_V:.2f} ft/s")
    return int_V

# Compute intersection velocities for positive and negative limit load factors with in the stall boundary
V_stall_pos_end = compute_intersection_velocity(stall_coeff, n_pos_limit[0])
V_stall_pos = np.linspace(0, V_stall_pos_end, 100)

V_stall_neg_end = compute_intersection_velocity(stall_coeff, n_neg_limit[0])
V_stall_neg = np.linspace(0, V_stall_neg_end, 100)

# Compute the corresponding n values at the intersection points
n_stall_pos_intersection = stall_coeff * V_stall_pos**2
n_stall_neg_intersection = -stall_coeff * V_stall_neg**2

plt.figure(figsize=(16,9))
plt.title('Maneuvering Envelope with Intersection Velocities')
plt.xlabel("V (ft/s)")
plt.ylabel("n (-)")
plt.plot(V_stall_pos, n_stall_pos_intersection  , label='Stall-Pos Intersection', linewidth=2, marker=None, markersize=8)
plt.plot(V_stall_neg, n_stall_neg_intersection  , label='Stall-Neg Intersection', linewidth=2, marker=None, markersize=8)
plt.grid(True)
plt.legend(loc='best')
plt.show()

# SECTION 5
# Extend the positive limit load line to the right of the intersection point
Vc = V_max * 0.8
V_pos_limit_extended = np.linspace(V_stall_pos_end, V_max, 100)
V_neg_limit_extended = np.linspace(V_stall_neg_end, Vc, 100)

n_pos_extended = n_pos_limit[0] * np.ones_like(V_pos_limit_extended)
n_neg_extended = n_neg_limit[0] * np.ones_like(V_neg_limit_extended)

# lastly, we can also plot the exceed speed lines
V_exceed = V_max 
# vertical line goes from negative limit load to positive limit load
n_exceed_line = np.linspace(0, n_pos_limit[0], 100)  
V_exceed_line = V_exceed * np.ones_like(n_exceed_line)

# SECTION 6
# Connect the end of the purple exceed line (V_exceed, 0) to the orange limit load neg line (Vc, -1)
V_connect = np.array([V_exceed, Vc])
n_connect = np.array([0, n_neg_limit[0]])

plt.figure(figsize=(16,9))
plt.title('Maneuvering Envelope with Extended Limit Load Lines')
plt.xlabel("V (ft/s)")
plt.ylabel("n (-)")
plt.plot(V_stall_pos, n_stall_pos_intersection, label='Stall-Pos', linewidth=2, marker=None, markersize=8, color='blue')
plt.plot(V_stall_neg, n_stall_neg_intersection, label='Stall-Neg', linewidth=2, marker=None, markersize=8, color='orange')
plt.plot(V_pos_limit_extended, n_pos_extended, label='Limit Load Pos', linewidth=2, marker=None, markersize=8, color='blue')
plt.plot(V_neg_limit_extended, n_neg_extended, label='Limit Load Neg', linewidth=2, marker=None, markersize=8, color='orange')
plt.plot(V_exceed_line, n_exceed_line, label='Exceed Speed', linewidth=2, color='purple')
plt.plot(V_connect, n_connect, label='Cruise-Dive Line', linewidth=2, marker=None, color='green')
plt.grid(True)
plt.legend(loc='best')
plt.show()

# SECTION 7 (OPTIONAL GUST)
def compute_k_g(MTOW, lift_slope, S_ref, c_bar):
    mu_g = 2*MTOW / (lift_slope * S_ref * c_bar)
    k_g = 0.88 * mu_g /(1 + mu_g)
    print(f"Computed k_g: {k_g:.3f} lb/ft^2 per unit CL")
    print(f"Computed mu_g: {mu_g:.3f}")
    return k_g

# SECTION 8 (OPTIONAL GUST)
b = ws["B9"].value            # ft, wingspan
c_bar = S_ref / b  # ft, mean aerodynamic chord
print(f"Computed mean aerodynamic chord: {c_bar:.3f} ft")

# SECTION 9 (OPTIONAL GUST)
lift_slope = ws["B121"].value     # 1/rad
k_g = compute_k_g(MTOW, lift_slope, S_ref, c_bar)

# SECTION 10 (OPTIONAL GUST)
def air_density_us_customary(altitude_ft):
    """Returns density in slugs/ft³ for given altitude in feet"""
    if altitude_ft < 0:
        raise ValueError("Altitude must be non-negative.")
    
    # Convert altitude to meters for the NASA formula
    altitude_m = altitude_ft * 0.3048
    
    # Troposphere
    if altitude_m <= 11000:
        T_K = 15.04 - 0.00649 * altitude_m + 273.15   # K
        p_Pa = 101290 * (T_K / 288.08) ** 5.256       # Pa
    
    # Lower stratosphere
    elif altitude_m <= 25000:
        T_K = -56.46 + 273.15                         # K
        p_Pa = 22650 * math.exp(1.73 - 0.000157 * altitude_m)   # Pa
    
    else:
        raise ValueError("This simplified NASA model here is set up only up to 25,000 m.")
    
    # Convert to US customary units
    # Pressure: Pa to lb/ft² (psf)
    p_psf = p_Pa * 0.0208854
    
    # Temperature: K to Rankine
    T_R = T_K * 1.8
    
    # Gas constant for air: 1716.5 ft·lb/(slug·°R)
    R = 1716.5
    
    # Density in slugs/ft³
    rho_slugs_ft3 = p_psf / (R * T_R)
    
    return rho_slugs_ft3

U_de_sl = 56.0          # ft/s (already correct)
U_de_15kft = 44.0       # ft/s (already correct)
U_de_50kft = 26.0       # ft/s (already correct)

print(f"Vertical gust velocity at sea level: {U_de_sl:.2f} ft/s")
print(f"Vertical gust velocity at 15,000 ft: {U_de_15kft:.2f} ft/s")
print(f"Vertical gust velocity at 50,000 ft: {U_de_50kft:.2f} ft/s")

U_de_sl_dive = U_de_sl * 0.5
U_de_15kft_dive = U_de_15kft * 0.5
U_de_50kft_dive = U_de_50kft * 0.5
print(f"Vertical gust velocity at sea level during dive: {U_de_sl_dive:.2f} ft/s")
print(f"Vertical gust velocity at 15,000 ft during dive: {U_de_15kft_dive:.2f} ft/s")
print(f"Vertical gust velocity at 50,000 ft during dive: {U_de_50kft_dive:.2f} ft/s")

rho_sl = air_density_us_customary(0)
print(f"Air density at sea level: {rho_sl:.6f} slugs/ft³")
rho_15kft = air_density_us_customary(15000)
print(f"Air density at 15,000 ft: {rho_15kft:.6f} slugs/ft³")
rho_50kft = air_density_us_customary(50000)
print(f"Air density at 50,000 ft: {rho_50kft:.6f} slugs/ft³")

# SECTION 11 (OPTIONAL GUST)
constant = lift_slope * k_g /(2 * MTOW / S_ref)
print(f"Constant term for gust boundary: {constant:.6f} (ft/s)^-2")

# SECTION 12 (OPTIONAL GUST)
rho_values = [rho_sl, rho_15kft, rho_50kft]
U_de_values = [U_de_sl, U_de_15kft, U_de_50kft]
altitude_labels = ["Sea Level", "15,000 ft", "50,000 ft"]

gust_boundary_list = []

for altitude, rho, U_de in zip(altitude_labels, rho_values, U_de_values):
    gust_boundary = constant * rho * U_de
    gust_boundary_list.append({
        "Altitude": altitude,
        "rho": rho,
        "U_de": U_de,
        "constant_term": gust_boundary
    })

gust_boundary_df = pd.DataFrame(gust_boundary_list)
gust_boundary_df

# SECTION 13 (OPTIONAL GUST)
# For dive condition, we can also compute the gust boundary values
U_de_dive_values = [U_de_sl_dive, U_de_15kft_dive, U_de_50kft_dive]

gust_boundary_dive_list = []

for altitude, rho, U_de_dive in zip(altitude_labels, rho_values, U_de_dive_values):
    gust_boundary_dive = constant * rho * U_de_dive
    gust_boundary_dive_list.append({
        "Altitude": altitude,
        "rho": rho,
        "U_de_dive": U_de_dive,
        "constant_term_dive": gust_boundary_dive
    })
gust_boundary_dive_df = pd.DataFrame(gust_boundary_dive_list)
gust_boundary_dive_df

# SECTION 14 (OPTIONAL GUST)
V_dive = V_max
V_gust = V_max * 0.8

def plot_gust_case(selected_altitude, V_gust, V_dive):
    # Plotting gust line from 0 to V_gust for cruise, and from 0 to V_dive for dive
    V_gust_array = np.linspace(0, V_gust, 200)
    V_dive_array = np.linspace(0, V_dive, 200)
    
    # Plotting a straight line connecting the cruise gust line endpoint at V_gust and the dive gust line endpoint at V_dive
    V_dive_connect_array = np.linspace(V_gust, V_dive, 200)

    plt.figure(figsize=(16,9))
    plt.title(f'Maneuvering Envelope with Gust Lines ({selected_altitude})')
    plt.xlabel("V (ft/s)")
    plt.ylabel("n (-)")

    # Existing maneuvering envelope
    plt.plot(V_stall_pos, n_stall_pos_intersection,
             label='Stall-Pos', linewidth=2, color='blue')
    plt.plot(V_stall_neg, n_stall_neg_intersection,
             label='Stall-Neg', linewidth=2, color='orange')
    plt.plot(V_pos_limit_extended, n_pos_extended,
             label='Limit Load Pos', linewidth=2, color='blue')
    plt.plot(V_neg_limit_extended, n_neg_extended,
             label='Limit Load Neg', linewidth=2, color='orange')
    plt.plot(V_exceed_line, n_exceed_line,
             label='Exceed Speed', linewidth=2, color='purple')
    plt.plot(V_connect, n_connect, label='Cruise-Dive Line', linewidth=2, marker=None, color='green')

    case_color = 'red'

    # Cruise
    row_cruise = gust_boundary_df[gust_boundary_df["Altitude"] == selected_altitude].iloc[0]
    gust_coeff = row_cruise["constant_term"]

    n_gust_pos = 1 + gust_coeff * V_gust_array
    n_gust_neg = 1 - gust_coeff * V_gust_array

    plt.plot(V_gust_array, n_gust_pos,
             color=case_color, linewidth=1.8, linestyle='-',
             label=f'Cruise Gust ({selected_altitude})')
    plt.plot(V_gust_array, n_gust_neg,
             color=case_color, linewidth=1.8, linestyle='-',
             label='_nolegend_')

    # Dive
    row_dive = gust_boundary_dive_df[gust_boundary_dive_df["Altitude"] == selected_altitude].iloc[0]
    gust_coeff_dive = row_dive["constant_term_dive"]

    # Original dive gust line (not connected to cruise gust line)
    n_gust_pos_dive_og = 1 + gust_coeff_dive * V_dive_array
    n_gust_neg_dive_og = 1 - gust_coeff_dive * V_dive_array

    plt.plot(V_dive_array, n_gust_pos_dive_og,
             color=case_color, linewidth=1.8, linestyle='--',
             label=f'Dive Gust ({selected_altitude})')
    plt.plot(V_dive_array, n_gust_neg_dive_og,
             color=case_color, linewidth=1.8, linestyle='--',
             label='_nolegend_')

    # endpoint at V_C from cruise gust line
    n_gust_pos_at_VC = 1 + gust_coeff * V_gust_array[-1]
    n_gust_neg_at_VC = 1 - gust_coeff * V_gust_array[-1]

    # endpoint at V_D from dive gust line
    n_gust_pos_at_VD = 1 + gust_coeff_dive * V_dive_array[-1]
    n_gust_neg_at_VD = 1 - gust_coeff_dive * V_dive_array[-1]

    # straight line between V_C and V_D
    n_gust_pos_dive = np.linspace(n_gust_pos_at_VC, n_gust_pos_at_VD, len(V_dive_array))
    n_gust_neg_dive = np.linspace(n_gust_neg_at_VC, n_gust_neg_at_VD, len(V_dive_array))

    plt.plot(V_dive_connect_array, n_gust_pos_dive,
             color=case_color, linewidth=1.8, linestyle='--',
             label=f'Dive Gust ({selected_altitude})')
    plt.plot(V_dive_connect_array, n_gust_neg_dive,
             color=case_color, linewidth=1.8, linestyle='--',
             label='_nolegend_')
    
    # Vertical line at V_dive between negative and positive dive gust endpoints
    plt.plot(
        [V_dive, V_dive],
        [n_gust_neg_at_VD, n_gust_pos_at_VD],
        color=case_color,
        linewidth=1.8,
        linestyle='--',
        label='_nolegend_'
    )
    plt.grid(True)
    plt.legend(loc='best')
    plt.show()


plot_gust_case("Sea Level", V_gust, V_dive)


# SECTION 15 (OPTIONAL GUST)
plot_gust_case("15,000 ft", V_gust, V_dive)

# CRITICAL V-n diagram points

# Calculate Vs (Stall speed at n=1)
Vs = np.sqrt(1.0 / stall_coeff)

# Calculate Va (Maneuvering speed = intersection of positive limit load w/ stall boundary)
n_pos_limit = 7
Va = np.sqrt(abs(n_pos_limit) / stall_coeff)

#Vc calcauted ealier

# Vd (Design dive speed)
Vd = V_max

# Vb (Design for Max Gust Intensity): intersection of stall-pos line and sea-level gust line

# Extract sea-level gust coefficient
gust_coeff_sea = gust_boundary_df[gust_boundary_df["Altitude"] == "Sea Level"]["constant_term"].iloc[0]

# Solve: stall_coeff*V^2 = 1 + gust_coeff_sea*V
a = stall_coeff
b = -gust_coeff_sea
c = -1

disc = b**2 - 4*a*c
if disc < 0:
    raise ValueError("No real intersection between gust line and stall line.")

Vb = (-b + np.sqrt(disc)) / (2*a)   # positive physical root
n_Vb = stall_coeff * Vb**2

print(f"Computed Vb (gust–stall intersection): {Vb:.2f} ft/s, n = {n_Vb:.2f}")


# Calculate Vh (Maximum level flight speed)
Vh = 0.9* Vd

# Create formatted table
print("\n" + "="*80)
print("FINAL V-n DIAGRAM CRITICAL VELOCITIES TABLE")
print("="*80)
print(f"{'Point':<10} {'Description':<45} {'Velocity (ft/s)':<18}")
print("-"*80)

# Table data
points = [
    ("Vs", "Stall speed", f"{Vs:.1f}"),
    ("Va", "Maneuvering Speed", f"{Va:.1f}"),
    ("Vb", "Design for Max Gust Intensity", f"{Vb:.1f}"),
    ("Vc", "Design Cruise Speed", f"{Vc:.1f}"),
    ("Vd", "Design Dive Speed", f"{Vd:.1f}"),
    ("Vh", "Maximum Level Flight Speed", f"{Vh:.1f}"),
]

for point in points:
    print(f"{point[0]:<10} {point[1]:<45} {point[2]:<18}")

print("="*80)

wb.save(r"C:\Users\Jaleal\Desktop\EAE 130B\Zephyr One Database.xlsx")