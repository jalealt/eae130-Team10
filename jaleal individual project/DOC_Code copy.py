import numpy as np

from openpyxl import load_workbook

file_path = r"C:\Users\Jaleal\Desktop\EAE 130B\Zephyr One Database.xlsx"
wb = load_workbook(file_path)
ws = wb["Database"]

## Inputs
MTOW = ws["B11"].value
R = ws["B20"].value
K = 2.75 # Regional
AF = 1 # Airline Factor 
b_year = 2006
t_year = 2026
fuel_weight = 20700 # Weight Code
pf = 6.84 # Approx Fuel density lb/gal
po = 7.5 # Approx oil density lb/gal
Pf = 4 # Approx Fuel Price per gallon
Po = 30 # Approx Oil Price per gallon
tb = 3 # Time of mission in hrs (assume 3 hrs)
CPI = 1.43 # 2012 to 2025
RL = 108 * CPI # Maintenance hourly rate (Used QC rate from Raymer)
C_unit = ws["B54"].value # Per Unit Cost
C_aircraft = ws["B55"].value # Flyaway Cost
C_engine = ws["B56"].value # Engine Cost
W_empty = ws["B29"].value # Aircraft empty weight
W_engine = ws["B31"].value # P135 Weight
To = ws["B12"].value # Thrust
n_useful = ws["B16"].value # Useful years (approx 25 yrs)
n_engine = ws["B14"].value # Number of engines

## Calculating CEF
b_CEF = 5.17053 + 0.104981 * (b_year - 2006)
t_CEF = 5.17053 + 0.104981 * (t_year - 2006)
CEF = t_CEF/b_CEF

## Calculatng Cash Operating Cost (COC)
C_crew = (482 + 0.59 * (MTOW/1000)) * CEF * tb
C_attendants = 0 # No attendants
C_fuel = 1.02 * fuel_weight * Pf / pf
oil_weight = 0.0125 * fuel_weight * (tb/100)
C_oil = 0.0125 * oil_weight * Po / po
C_airport = 0 # No landing fees since not commercial
C_navigation = 0.5 * CEF * (1.852 * R / tb) * np.sqrt(0.0004535937 * MTOW / 50)

# Airframe Maintenance
WA = W_empty - W_engine # airframe weight
C_airframe = C_aircraft - C_engine # airframe cost
CML = 1.03 * (3 + 0.067 * WA / 1000) * RL
CMM = (1.03 * 30 * CEF) + (0.79 * 10**-5 * C_airframe)
C_airframe_maintenance = (CML + CMM) * tb

# Engine Maintenance
CML_2 = (0.645 + 0.05 * To * 10**-4) * (0.566 + 0.434 / tb) * RL
CMM_2 = (25 + 18 * To * 10**-4) * (0.62 + 0.38 / tb) * CEF
C_engine_maintenance = n_engine*(CML_2 + CMM_2) * tb

COC = C_crew + C_attendants + C_fuel + C_oil + C_airport + C_navigation + C_airframe_maintenance + C_engine_maintenance
print(f"COC: ${COC:,}")

## Calculating FOC
U_annual = 1.5 * 10**3 * (3.4546 * tb + 2.994 - (12.289 * tb **2 - 5.6626 * tb - 8.964) ** 0.5)
C_insurance = (0.02 * C_aircraft * tb) / U_annual
C_depreciation = C_unit * (1 - 0.1) * tb / (n_useful * U_annual)

# These should be in terms of DOC
C_registration = (0.001 + MTOW * 10 **-8)
C_fin = 0.07

FOC = C_insurance + C_depreciation
print(f"FOC: ${FOC:,}")

## DOC 
DOC = (COC + FOC) / (1 - C_fin - C_registration)
print(f"DOC: ${DOC:,}")

ws["F53"].value = COC
ws["F54"].value = FOC
ws["F55"].value = DOC

wb.save(r"C:\Users\Jaleal\Desktop\EAE 130B\Zephyr One Database.xlsx")