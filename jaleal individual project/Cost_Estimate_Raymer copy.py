from openpyxl import load_workbook

file_path = r"C:\Users\Jaleal\Desktop\EAE 130B\Zephyr One Database.xlsx"
wb = load_workbook(file_path)
ws = wb["Database"]

## INPUTS ##
W = ws["B29"].value # empty weight (lbs)
W_av = ws["B34"].value # lbs
V = ws["B94"].value # Max velocity (kt)
Q = ws["B51"].value # Number to be produced in five years
FTA = 3 # Number of flight-test aircraft
N = ws["B14"].value # Number of engines
K = 520 # Change to K = 436 for Turbojets OR K = 520 for Turbofans
N_eng = Q * N # Total production quantity times number of engines per aircraft
T_max = ws["B12"].value # Engine maximum thrust (lb)
M_max = ws["B17"].value # Engine maximum Mach number
T_turb_inlet = 4000 # Turbine inlet tempearature (approx 3700 - 4100 rankine)
CPI = ws["B52"].value # 2012 to 2025

# 0. Engineering Cost
# 1. Tooling Cost
# 2. Manufacturing Cost
# 3. Quality Check Cost
# 4. Development Support Cost
# 5. Flight Test Cost 
# 6. Manufacturing Materials Cost 

# Hourly rates from Raymer
R_eng = 115 
R_tool = 118
R_qc = 108
R_mfg = 98

R = [R_eng, R_tool, R_mfg, R_qc, 1, 1, 1]

# Coefficient at the start of equation
coeff = [4.86, 5.99, 7.37, 0.076, 91.3, 2498, 22.1]

# Weight Exponent
W_exp = [0.777, 0.777, 0.82, 0, 0.630, 0.325, 0.921]

# Velocity Exponent
V_exp = [0.894, 0.696, 0.484, 0, 1.3, 0.822, 0.621]

# Q Exponent
Q_exp  = [0.163, 0.263, 0.641, 0, 0, 0, 0.799]

# FTA Exponent
FTA_exp = [0, 0, 0, 0, 0, 1.21, 0]

# Total Cost for RDT&E
rdte = [0,4,5] # Numbers corresponding to the CER categories that are considered RDT&E
C_rdte = [ ]

for i in rdte:
    C_category = coeff[i] * (W**W_exp[i]) * (V**V_exp[i]) * (Q**Q_exp[i]) * (FTA**FTA_exp[i]) * R[i] *CPI
    C_rdte.append(C_category)

C_rdte_tot = sum(C_rdte)

# Total production cost 
prod = [1, 2, 3, 6] # Numbers corresponding to the CER categories that are considered production
C_prod = [ ]

for i in prod:
    if i == 3:
        y = 2
        mfg_hours = coeff[y] * (W**W_exp[y]) * (V**V_exp[y]) * (Q**Q_exp[y]) * (FTA**FTA_exp[y])
        C_category = coeff[i] * R[i] * mfg_hours* CPI
        C_prod.append(C_category)
    else:
        C_category = coeff[i] * (W**W_exp[i]) * (V**V_exp[i]) * (Q**Q_exp[i]) * (FTA**FTA_exp[i]) * R[i] *CPI
        C_prod.append(C_category)

C_prod_tot = sum(C_prod)

# Avionics Cost
CPI_av = 2.96
avionics_cost = 2000 * W_av * CPI_av

# Total Engine cost
# C_engine = 3112 *(0.043 * T_max + 243.25 * M_max + 0.969 *T_turb_inlet - 2228) * CPI
C_engine = 20400000 # Cost of P135 P&W 100

# Total price
C_tot = C_prod_tot + C_rdte_tot
C_unit = ((C_rdte_tot+C_prod_tot)/Q)
C_unit_2 = C_unit + C_engine + avionics_cost

C_flyaway = (C_prod_tot/Q) + C_engine + avionics_cost

print(f"Total RDT&E Cost: ${C_rdte_tot:,}")
print(f"Total Production Cost: ${C_prod_tot:,}")
print(f"Total Cost of RDT&E and Production Combined: ${C_tot:,}\n") 
print(f"Total Engine Cost: ${C_engine:,}")
print(f"Per Unit Cost w/o Engine: ${C_unit:,}")
print(f"Per Unit Cost w Engine: ${C_unit_2:,}")
print(f"Total Flyaway Cost: ${C_flyaway:,}")

ws["F51"].value = C_rdte_tot
ws["F52"].value = C_prod_tot
ws["B54"].value = C_unit_2
ws["B55"].value = C_flyaway

wb.save(r"C:\Users\Jaleal\Desktop\EAE 130B\Zephyr One Database.xlsx")